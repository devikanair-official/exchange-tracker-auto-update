import json
from pathlib import Path
from crewai.tools import BaseTool
from pydantic import BaseModel, Field
from openpyxl import load_workbook

# ── Columns the agent is ALLOWED to update ────────────────────────────────────
SAFE_COLUMNS = {
    "Platform ADV",
    "Our Volume",
    "Migration Status",
    "ETA",
}

class WritebackInput(BaseModel):
    tracker_path: str = Field(..., description="Absolute or relative path to the Excel tracker file.")
    updates: str = Field(
        ...,
        description=(
            "JSON string — a list of row update objects. Each object must have:\n"
            "  'exchange'  : exchange name to match (column A)\n"
            "  'field'     : column header to update (must be in safe list)\n"
            "  'value'     : new value to write\n"
            "Example: [{\"exchange\":\"Binance\",\"field\":\"Migration Status\",\"value\":\"In Progress\"}]"
        ),
    )

class ExcelWritebackTool(BaseTool):
    name: str = "Excel Writeback Tool"
    description: str = (
        "Safely updates specific fields in the Exchange Migration Tracker Excel file. "
        f"Only these columns can be modified: {', '.join(SAFE_COLUMNS)}. "
        "All other columns are protected and will not be touched."
    )
    args_schema: type[BaseModel] = WritebackInput

    def _run(self, tracker_path: str, updates: str) -> str:
        path = Path(tracker_path)
        if not path.exists():
            return f"ERROR: Tracker file not found at '{tracker_path}'"

        try:
            update_list = json.loads(updates)
        except json.JSONDecodeError as e:
            return f"ERROR: Could not parse updates JSON — {e}"

        wb = load_workbook(path)
        ws = wb.active

        # Build header → column index map
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}

        results = []
        for item in update_list:
            exchange = item.get("exchange")
            field    = item.get("field")
            value    = item.get("value")

            if field not in SAFE_COLUMNS:
                results.append(f"SKIPPED '{field}' for {exchange} — not in safe column list.")
                continue

            if field not in headers:
                results.append(f"SKIPPED '{field}' for {exchange} — column not found in tracker.")
                continue

            col_idx = headers[field]
            matched = False
            for row in ws.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).strip().lower() == str(exchange).strip().lower():
                    ws.cell(row=row[0].row, column=col_idx, value=value)
                    results.append(f"UPDATED {exchange} → {field} = {value}")
                    matched = True
                    break

            if not matched:
                results.append(f"NOT FOUND: '{exchange}' — no matching row in tracker.")

        wb.save(path)
        return "\n".join(results)
